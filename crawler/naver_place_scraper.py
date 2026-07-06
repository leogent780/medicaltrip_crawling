"""네이버 검색결과(플레이스 탭)와 플레이스 상세를 순수 HTTP 요청으로 조회하는 크롤러.

Playwright 같은 브라우저 자동화 없이 requests로만 동작한다. 지도 화면(map.naver.com)을
직접 열고 클릭하는 방식보다 "자동화 브라우저" 신호가 훨씬 적어서 캡차가 덜 뜬다
(피부과 업체 수집 때 이 방식으로 성공한 사례 기반).

주의:
- search.naver.com 검색결과 페이지에 내장된 JSON을 정규식으로 그대로 뽑는 방식이라
  공식 계약이 아니다. 네이버가 마크업/필드 구성을 바꾸면 깨질 수 있다.
- map.naver.com/p/api/place/summary, m.place.naver.com/place/* 도 마찬가지로 비공식
  내부 엔드포인트라 언제든 바뀔 수 있다.
- 개인 리서치용 소규모 수집을 전제로 만들었다. SEARCH_DELAY/DETAIL_DELAY를 임의로
  줄이지 말 것.
"""

import argparse
import re
import time
from dataclasses import dataclass
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SEARCH_DELAY = 1.0
DETAIL_DELAY = 0.5
BLOG_DELAY = 0.5

REGION_ALIASES = {
    "명동": ["명동", "을지로입구", "충무로"],
    "성수": ["성수", "성수동", "뚝섬"],
    "홍대": ["홍대", "연남동", "합정"],
    "강남": ["강남역", "역삼동", "논현동"],
    "압구정": ["압구정", "압구정로데오"],
    "신사": ["신사동", "가로수길"],
    "용산": ["용산", "이태원", "한남동"],
}

INSTAGRAM_PATTERNS = [
    re.compile(r"instagram\.com/([A-Za-z0-9._]{2,30})"),
    re.compile(r'"instagram"\s*:\s*"([^"]+)"'),
    re.compile(r'instaId["\s:]+([A-Za-z0-9._]{2,30})'),
]
# 블로그 후기 글은 링크 없이 "@계정명"만 텍스트로 쓰는 경우가 많아 보조로 사용.
# 이메일 주소를 계정으로 오인하기 쉬워서 흔한 메일 도메인은 걸러낸다.
BLOG_INSTAGRAM_RE = re.compile(r"(?<![\w.])@([A-Za-z0-9](?:[A-Za-z0-9._]{1,28})[A-Za-z0-9])")
EMAIL_LIKE_DOMAINS = {
    "gmail.com", "naver.com", "daum.net", "hanmail.net", "nate.com",
    "kakao.com", "hotmail.com", "yahoo.com", "icloud.com", "outlook.com",
}
KAKAO_RE = re.compile(r"pf\.kakao\.com/[A-Za-z0-9_]+")
TAG_RE = re.compile(r"<[^>]+>")
IGNORE_INSTA_HANDLES = {"p", "reel", "explore", "stories", "accounts", ""}


def make_session():
    session = requests.Session()
    session.headers.update({
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
        "Accept-Language": "ko-KR,ko;q=0.9",
    })
    return session


@dataclass
class Place:
    id: str
    region: str
    name: str = ""
    address: str = ""
    phone: str = ""
    instagram: str = ""
    kakao_channel: str = ""
    first_visit_event: str = ""
    matched_keywords: str = ""


def add_matched_keyword(place, keyword):
    existing = [k for k in place.matched_keywords.split(",") if k]
    if keyword not in existing:
        existing.append(keyword)
    place.matched_keywords = ",".join(existing)


def build_queries(region, keyword):
    aliases = REGION_ALIASES.get(region, [region])
    return [f"서울 {alias} {keyword}" for alias in aliases]


MAX_SEARCH_PAGES = 5


def _parse_search_page(text):
    ids = re.findall(r'"id"\s*:\s*"(\d{7,})"', text)
    names = re.findall(r'"name"\s*:\s*"([^"]{2,30})"', text)
    addresses = re.findall(r'"roadAddress"\s*:\s*"([^"]*)"', text)
    phones = re.findall(r'"phone"\s*:\s*"([^"]*)"', text)
    return [
        {"id": pid,
         "name": names[i] if i < len(names) else "",
         "address": addresses[i] if i < len(addresses) else "",
         "phone": phones[i] if i < len(phones) else ""}
        for i, pid in enumerate(ids)
    ]


def search_places(session, query, region, max_results=70, debug_dir=None, log=print):
    """search.naver.com의 place 탭은 한 번에 미리보기 몇 개만 보여주는 것 같아서,
    start 페이지네이션을 시도한다. 페이지가 새 업체를 안 주면 바로 멈춘다
    (네이버가 start를 무시해도 안전하게 종료됨)."""
    url = "https://search.naver.com/search.naver"
    safe_name = re.sub(r"[^0-9A-Za-z가-힣]+", "_", query)
    found = {}

    start = 1
    for page in range(MAX_SEARCH_PAGES):
        if len(found) >= max_results:
            break
        params = {"where": "place", "query": query, "display": 70, "start": start}
        try:
            resp = session.get(url, params=params, timeout=15)
        except requests.RequestException as e:
            log(f"  [오류] '{query}' 검색 실패: {e}")
            break

        text = resp.text
        if debug_dir:
            (debug_dir / f"search_{safe_name}_p{page + 1}.html").write_text(text, encoding="utf-8")

        page_items = _parse_search_page(text)
        new_count = 0
        for item in page_items:
            if item["id"] not in found:
                found[item["id"]] = item
                new_count += 1

        if page == 0:
            log(f"    (검색 1페이지: {len(page_items)}개 파싱됨)")
        if new_count == 0:
            break

        start += len(page_items) if page_items else 15
        time.sleep(SEARCH_DELAY)

    return [
        Place(id=item["id"], region=region, name=item["name"],
              address=item["address"], phone=item["phone"])
        for item in found.values()
    ]


def _extract_instagram(text, loose=False):
    for pattern in INSTAGRAM_PATTERNS:
        m = pattern.search(text)
        if m:
            handle = m.group(1).strip("/").strip('"')
            if handle.lower() not in IGNORE_INSTA_HANDLES:
                return f"https://www.instagram.com/{handle}/"
    if loose:
        m = BLOG_INSTAGRAM_RE.search(text)
        if m:
            handle = m.group(1)
            if handle.lower() not in IGNORE_INSTA_HANDLES and handle.lower() not in EMAIL_LIKE_DOMAINS:
                return f"https://www.instagram.com/{handle}/"
    return ""


def _extract_first_visit(text):
    idx = text.find("첫방문")
    if idx == -1:
        return ""
    start = text.rfind(" ", 0, max(0, idx - 15)) + 1
    return TAG_RE.sub(" ", text[start: idx + 60]).strip()


def enrich_place(session, place, debug_dir=None, log=print):
    detail_urls = [
        f"https://map.naver.com/p/api/place/summary/{place.id}",
        f"https://m.place.naver.com/place/{place.id}/home",
    ]
    combined_text = ""
    for url in detail_urls:
        try:
            resp = session.get(url, timeout=10)
            combined_text += resp.text
        except requests.RequestException:
            continue

    if debug_dir:
        (debug_dir / f"{place.id}_detail.txt").write_text(combined_text, encoding="utf-8")

    place.instagram = _extract_instagram(combined_text)

    m = KAKAO_RE.search(combined_text)
    if m:
        place.kakao_channel = m.group(0)

    try:
        resp = session.get(f"https://m.place.naver.com/place/{place.id}/feed", timeout=10)
        if debug_dir:
            (debug_dir / f"{place.id}_feed.txt").write_text(resp.text, encoding="utf-8")
        place.first_visit_event = _extract_first_visit(resp.text)
    except requests.RequestException:
        pass

    # 플레이스 등록정보에 인스타/첫방문 이벤트가 없으면 블로그 후기에서 보조로 찾는다.
    if not place.instagram or not place.first_visit_event:
        find_via_blog(session, place, debug_dir=debug_dir, log=log)

    return place


def find_via_blog(session, place, debug_dir=None, log=print):
    query = f"{place.name} {place.region} 인스타그램 후기"
    try:
        resp = session.get("https://search.naver.com/search.naver",
                            params={"where": "post", "query": query}, timeout=10)
        time.sleep(BLOG_DELAY)
    except requests.RequestException as e:
        log(f"    [WARN] 블로그 보조검색 실패: {place.name} - {e}")
        return

    text = TAG_RE.sub(" ", resp.text)
    if debug_dir:
        (debug_dir / f"{place.id}_blog.txt").write_text(text, encoding="utf-8")

    if not place.instagram:
        place.instagram = _extract_instagram(text, loose=True)
    if not place.first_visit_event:
        place.first_visit_event = _extract_first_visit(text)


def crawl(regions, keywords, max_per_region, log=print, debug_dir=None):
    if isinstance(keywords, str):
        keywords = [k.strip() for k in keywords.split(",") if k.strip()]
    if not keywords:
        keywords = ["에스테틱"]

    session = make_session()
    all_places = {}

    for region in regions:
        for keyword in keywords:
            log(f"[INFO] === {region} / {keyword} 검색 시작 ===")
            found_ids_this_keyword = set()
            for query in build_queries(region, keyword):
                log(f"[INFO] 검색: {query}")
                found = search_places(session, query, region, debug_dir=debug_dir, log=log)
                new = 0
                for p in found:
                    found_ids_this_keyword.add(p.id)
                    if p.id not in all_places:
                        all_places[p.id] = p
                        new += 1
                    add_matched_keyword(all_places[p.id], keyword)
                log(f"  -> {new}개 신규 (이 키워드 누적 {len(found_ids_this_keyword)}개)")
                time.sleep(SEARCH_DELAY)
                if len(found_ids_this_keyword) >= max_per_region:
                    break

    log(f"\n[INFO] 총 {len(all_places)}곳 발견. 상세정보(인스타/카톡/첫방문이벤트) 조회 중...")
    for place in all_places.values():
        enrich_place(session, place, debug_dir=debug_dir, log=log)
        log(f"    - {place.name} | {place.address} | IG:{place.instagram or '-'}")
        time.sleep(DETAIL_DELAY)

    return list(all_places.values())


def save_excel(places, output_path, log=print):
    wb = Workbook()
    ws = wb.active
    ws.title = "에스테틱 리스트"
    headers = ["지역", "업체명", "검색키워드", "주소", "전화번호", "인스타그램",
               "카카오톡채널", "첫방문이벤트", "네이버플레이스ID"]
    ws.append(headers)

    seen = set()
    for p in places:
        key = (p.name, p.address)
        if key in seen:
            continue
        seen.add(key)
        ws.append([p.region, p.name, p.matched_keywords, p.address, p.phone, p.instagram,
                   p.kakao_channel, p.first_visit_event, p.id])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    wb.save(output_path)
    log(f"[INFO] 엑셀 저장 완료: {output_path} (총 {len(seen)}곳)")


def main():
    parser = argparse.ArgumentParser(description="네이버 검색/플레이스 HTTP 기반 에스테틱 크롤러")
    parser.add_argument("--regions", nargs="+",
                         default=["명동", "성수", "홍대", "강남", "압구정", "신사", "용산"])
    parser.add_argument("--keywords", nargs="+", default=["에스테틱"],
                         help="예: --keywords 에스테틱 스파 마사지 헤어 메이크업")
    parser.add_argument("--max-per-region", type=int, default=15)
    parser.add_argument("--output", default="seoul_esthetic_list.xlsx")
    parser.add_argument("--debug", action="store_true",
                         help="업체별 상세/피드/블로그 원문 응답을 debug/ 에 저장")
    args = parser.parse_args()

    debug_dir = None
    if args.debug:
        debug_dir = Path("debug")
        debug_dir.mkdir(exist_ok=True)

    places = crawl(args.regions, args.keywords, args.max_per_region, debug_dir=debug_dir)
    save_excel(places, args.output)


if __name__ == "__main__":
    main()
