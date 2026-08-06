"""네이버 지도에서 '지역+키워드'로 업체를 검색해 엑셀로 정리하는 크롤러.

사용법:
    pip install -r requirements.txt
    playwright install chromium
    python naver_esthetic_crawler.py --regions 명동 성수 홍대 --keyword 에스테틱 --max-per-region 30

네이버 지도 화면의 class 이름은 해시값이라 수시로 바뀐다. 이 스크립트는 오래 유지된
구조적 id(#searchIframe, #entryIframe, #_pcmap_list_scroll_container)와 텍스트 휴리스틱에
의존하므로 셀렉터가 언젠가 깨질 수 있다. --debug 를 켜면 단계별 스크린샷을 debug/ 에 남긴다.
"""

import argparse
import random
import re
import time
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PWTimeout
from playwright.sync_api import sync_playwright

# 너무 짧게 잡으면 네이버 쪽에서 접근을 막을 수 있으니 임의로 줄이지 말 것.
REQUEST_DELAY = (1.5, 3.0)
NAV_TIMEOUT = 15000

INSTAGRAM_RE = re.compile(r"instagram\.com/([A-Za-z0-9_.]+)")
KAKAO_RE = re.compile(r"pf\.kakao\.com/[A-Za-z0-9_]+")
ADDRESS_RE = re.compile(r"[가-힣]+(?:로|길)\s?\d+[^\n,]*")
PHONE_RE = re.compile(r"0\d{1,2}-?\d{3,4}-?\d{4}")


DEFAULT_REGIONS = {
    "서울": ["명동", "성수", "홍대", "강남", "압구정", "신사", "용산"],
    "부산": ["서면", "해운대", "동래", "남포동", "부산대", "연산동", "사상", "광안리"],
}


@dataclass
class Place:
    region: str
    name: str = ""
    category: str = ""
    address: str = ""
    phone: str = ""
    homepage: str = ""
    instagram: str = ""
    kakao_channel: str = ""
    first_visit_event: str = ""
    naver_url: str = ""


def polite_sleep():
    time.sleep(random.uniform(*REQUEST_DELAY))


def get_search_frame(page):
    frame_el = page.wait_for_selector("iframe#searchIframe", timeout=NAV_TIMEOUT)
    return frame_el.content_frame()


def get_entry_frame(page):
    frame_el = page.wait_for_selector("iframe#entryIframe", timeout=NAV_TIMEOUT)
    return frame_el.content_frame()


def collect_list_items(frame, max_count, debug_dir=None, tag=""):
    """검색 결과 목록을 스크롤하며 li 항목을 모은다. 업체명은 각 li의 첫 텍스트 줄로 추정한다."""
    scroll_sel = "#_pcmap_list_scroll_container"
    try:
        frame.wait_for_selector(scroll_sel, timeout=NAV_TIMEOUT)
    except PWTimeout:
        if debug_dir:
            frame.page.screenshot(path=str(debug_dir / f"{tag}_no_list.png"))
        return []

    seen_count = 0
    stable_rounds = 0
    while stable_rounds < 3:
        items = frame.query_selector_all(f"{scroll_sel} li")
        if len(items) >= max_count or len(items) == seen_count:
            stable_rounds += 1
        else:
            stable_rounds = 0
        seen_count = len(items)
        if len(items) >= max_count:
            break
        frame.eval_on_selector(scroll_sel, "(el) => el.scrollBy(0, el.scrollHeight)")
        time.sleep(1.0)

    items = frame.query_selector_all(f"{scroll_sel} li")[:max_count]
    results = []
    for li in items:
        text = (li.inner_text() or "").strip()
        if not text:
            continue
        lines = [line.strip() for line in text.split("\n") if line.strip()]
        if not lines:
            continue
        link_el = li.query_selector("a")
        results.append({
            "name": lines[0],
            "category": lines[1] if len(lines) > 1 else "",
            "el": li,
            "link": link_el,
        })

    if debug_dir:
        frame.page.screenshot(path=str(debug_dir / f"{tag}_list.png"))

    return results


def enrich_from_entry(frame, place, debug_dir=None, tag=""):
    try:
        frame.wait_for_selector("a[href*='http']", timeout=NAV_TIMEOUT)
    except PWTimeout:
        pass

    full_text = frame.locator("body").inner_text()

    addr_match = ADDRESS_RE.search(full_text)
    if addr_match:
        place.address = addr_match.group(0).strip()

    phone_match = PHONE_RE.search(full_text)
    if phone_match:
        place.phone = phone_match.group(0)

    for a in frame.query_selector_all("a[href]"):
        href = a.get_attribute("href") or ""
        ig_match = INSTAGRAM_RE.search(href)
        if ig_match and not place.instagram:
            place.instagram = "@" + ig_match.group(1)
        if KAKAO_RE.search(href) and not place.kakao_channel:
            place.kakao_channel = href
        if (href.startswith("http") and not place.homepage
                and "naver.com" not in href and "instagram.com" not in href
                and "kakao.com" not in href):
            place.homepage = href

    # 업체가 올린 소식(피드) 게시글에 "첫방문" 문구가 있는지만 확인한다.
    # 블로그 후기 기반 이벤트 탐지는 범위 밖(정확도가 낮아 별도 검토가 필요).
    try:
        feed_tab = frame.get_by_text("소식", exact=True).first
        if feed_tab:
            feed_tab.click(timeout=5000)
            frame.wait_for_timeout(1500)
            feed_text = frame.locator("body").inner_text()
            idx = feed_text.find("첫방문")
            if idx != -1:
                place.first_visit_event = feed_text[max(0, idx - 20): idx + 80].replace("\n", " ")
    except Exception:
        pass

    if debug_dir:
        frame.page.screenshot(path=str(debug_dir / f"{tag}_entry.png"))


def crawl_region(page, city, region, keyword, max_count, debug_dir=None, log=print):
    query = f"{city} {region} {keyword}"
    log(f"[INFO] 검색: {query}")
    page.goto(f"https://map.naver.com/p/search/{query}", timeout=NAV_TIMEOUT)
    polite_sleep()

    frame = get_search_frame(page)
    items = collect_list_items(frame, max_count, debug_dir=debug_dir, tag=f"{region}_search")
    log(f"[INFO]  -> {len(items)}개 항목 발견")

    places = []
    for idx, item in enumerate(items):
        name = item["name"]
        if not name:
            continue
        place = Place(region=region, name=name, category=item["category"])
        try:
            (item["link"] or item["el"]).click()
            page.wait_for_timeout(1200)
            entry_frame = get_entry_frame(page)
            place.naver_url = entry_frame.url
            enrich_from_entry(entry_frame, place, debug_dir=debug_dir, tag=f"{region}_{idx}_{name}")
        except Exception as e:
            log(f"    [WARN] '{name}' 상세정보 수집 실패: {e}")
        places.append(place)
        log(f"    - {place.name} | {place.address} | IG:{place.instagram or '-'}")
        polite_sleep()

    return places


def save_excel(places, output_path, log=print):
    wb = Workbook()
    ws = wb.active
    ws.title = "업체 리스트"
    headers = ["지역", "업체명", "카테고리", "주소", "전화번호", "인스타그램",
               "카카오톡채널", "홈페이지", "첫방문이벤트", "네이버플레이스"]
    ws.append(headers)

    seen = set()
    for p in places:
        key = (p.name, p.address)
        if key in seen:
            continue
        seen.add(key)
        ws.append([p.region, p.name, p.category, p.address, p.phone, p.instagram,
                   p.kakao_channel, p.homepage, p.first_visit_event, p.naver_url])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    wb.save(output_path)
    log(f"[INFO] 엑셀 저장 완료: {output_path} (총 {len(seen)}곳)")


def main():
    parser = argparse.ArgumentParser(description="네이버 지도 업체 크롤러")
    parser.add_argument("--city", default="서울", choices=sorted(DEFAULT_REGIONS.keys()),
                         help="기본 지역 목록을 고를 도시 (예: 서울, 부산)")
    parser.add_argument("--regions", nargs="+", default=None,
                         help="생략하면 --city의 기본 지역 목록을 사용한다")
    parser.add_argument("--keyword", default="에스테틱")
    parser.add_argument("--max-per-region", type=int, default=30)
    parser.add_argument("--output", default=None)
    parser.add_argument("--headless", action="store_true", help="브라우저 창 없이 실행")
    parser.add_argument("--debug", action="store_true", help="단계별 스크린샷을 debug/ 에 저장")
    args = parser.parse_args()

    regions = args.regions or DEFAULT_REGIONS.get(args.city, [args.city])
    output = args.output or f"{args.city}_{args.keyword}_list.xlsx"

    debug_dir = None
    if args.debug:
        debug_dir = Path("debug")
        debug_dir.mkdir(exist_ok=True)

    all_places = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless)
        page = browser.new_page()
        for region in regions:
            try:
                places = crawl_region(page, args.city, region, args.keyword, args.max_per_region, debug_dir=debug_dir)
                all_places.extend(places)
            except Exception as e:
                print(f"[ERROR] '{region}' 처리 중 오류: {e}")
        browser.close()

    save_excel(all_places, output)


if __name__ == "__main__":
    main()
