"""네이버 오픈 API(지역검색 + 블로그검색)로 업체 정보를 모으는 크롤러.

네이버 지도 화면(map.naver.com)을 자동화하지 않고 공식 REST API만 호출하기 때문에
캡차가 뜨지 않는다. 사용 전 준비:

1. https://developers.naver.com/apps 에서 애플리케이션 등록, 사용 API에 "검색" 체크
2. 발급받은 Client ID / Client Secret 을 crawler/.env 에 아래처럼 저장
       NAVER_CLIENT_ID=발급받은값
       NAVER_CLIENT_SECRET=발급받은값
   (.env.example 참고. .env 는 git에 올라가지 않는다.)

주의: 네이버 지역검색 API는 한 쿼리당 최대 5건만 돌려주고 페이지네이션도 없다
(공식 스펙 제약). 그래서 지역 하나당 여러 개의 쿼리 변형(동네 별칭 x 키워드 조합)을
돌려서 결과를 모으고 중복 제거하는 방식으로 개수를 늘린다.
"""

import argparse
import html
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

LOCAL_URL = "https://openapi.naver.com/v1/search/local.json"
BLOG_URL = "https://openapi.naver.com/v1/search/blog.json"

TAG_RE = re.compile(r"<.*?>")
INSTAGRAM_RE = re.compile(r"(?:instagram\.com/|(?<!\w)@)([A-Za-z0-9_.]{2,30})")
KAKAO_RE = re.compile(r"pf\.kakao\.com/[A-Za-z0-9_]+")

# 초당 호출 제한에 여유를 두기 위한 최소 지연. 공식 API라 캡차는 없지만 예의상 유지.
API_DELAY = 0.3

KEYWORD_VARIANTS_BY_CITY = {
    "서울": ["피부관리", "왁싱", "스킨케어"],
    "부산": ["피부과", "피부관리", "레이저클리닉"],
}
REGION_ALIASES = {
    "서울": {
        "명동": ["명동", "을지로입구", "충무로"],
        "성수": ["성수", "성수동", "뚝섬"],
        "홍대": ["홍대", "연남동", "합정"],
        "강남": ["강남역", "역삼동", "논현동"],
        "압구정": ["압구정", "압구정로데오"],
        "신사": ["신사동", "가로수길"],
        "용산": ["용산", "이태원", "한남동"],
    },
    "부산": {
        "서면": ["서면", "부전동", "전포동"],
        "해운대": ["해운대", "우동", "센텀시티"],
        "동래": ["동래", "명륜동", "온천동"],
        "남포동": ["남포동", "광복동", "중앙동"],
        "부산대": ["부산대", "장전동", "온천장"],
        "연산동": ["연산동", "거제동"],
        "사상": ["사상", "덕포동"],
        "광안리": ["광안리", "남천동"],
    },
}

DEFAULT_REGIONS = {
    "서울": ["명동", "성수", "홍대", "강남", "압구정", "신사", "용산"],
    "부산": ["서면", "해운대", "동래", "남포동", "부산대", "연산동", "사상", "광안리"],
}


def load_env():
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        os.environ.setdefault(key.strip(), value.strip())


def get_credentials():
    load_env()
    client_id = os.environ.get("NAVER_CLIENT_ID")
    client_secret = os.environ.get("NAVER_CLIENT_SECRET")
    if not client_id or not client_secret:
        raise RuntimeError(
            "NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 이 설정되어 있지 않다. "
            "crawler/.env.example 을 crawler/.env 로 복사한 뒤 발급받은 키를 채워라."
        )
    return client_id, client_secret


def strip_tags(text):
    return html.unescape(TAG_RE.sub("", text or "")).strip()


@dataclass
class Business:
    region: str
    name: str
    category: str = ""
    address: str = ""
    phone: str = ""
    instagram: str = ""
    kakao_channel: str = ""
    homepage: str = ""
    first_visit_event: str = ""
    source_blog: str = ""


def search_local(session, headers, query, log=print):
    params = {"query": query, "display": 5}
    resp = session.get(LOCAL_URL, params=params, headers=headers, timeout=10)
    time.sleep(API_DELAY)
    if resp.status_code != 200:
        log(f"[WARN] 지역검색 실패({resp.status_code}): {query} - {resp.text[:200]}")
        return []
    items = resp.json().get("items", [])
    return [{
        "name": strip_tags(item.get("title", "")),
        "category": item.get("category", ""),
        "address": item.get("roadAddress") or item.get("address") or "",
        "phone": item.get("telephone", ""),
    } for item in items]


def search_blog(session, headers, query, log=print):
    params = {"query": query, "display": 5, "sort": "sim"}
    resp = session.get(BLOG_URL, params=params, headers=headers, timeout=10)
    time.sleep(API_DELAY)
    if resp.status_code != 200:
        log(f"[WARN] 블로그검색 실패({resp.status_code}): {query} - {resp.text[:200]}")
        return []
    items = resp.json().get("items", [])
    return [{
        "title": strip_tags(it.get("title", "")),
        "desc": strip_tags(it.get("description", "")),
        "link": it.get("link", ""),
    } for it in items]


def fetch_blog_text(session, url, log=print):
    """blog.naver.com 링크를 모바일 버전으로 바꿔서 iframe 없이 본문 텍스트를 바로 받는다."""
    m_url = url.replace("blog.naver.com", "m.blog.naver.com")
    try:
        resp = session.get(m_url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        time.sleep(API_DELAY)
        if resp.status_code != 200:
            return ""
        return html.unescape(TAG_RE.sub(" ", resp.text))
    except requests.RequestException as e:
        log(f"[WARN] 블로그 본문 조회 실패: {url} - {e}")
        return ""


def collect_region(session, headers, city, region, keyword, max_count, log=print):
    seen = {}
    aliases = REGION_ALIASES.get(city, {}).get(region, [region])
    variants = KEYWORD_VARIANTS_BY_CITY.get(city, [])
    keywords = [keyword] + [k for k in variants if k != keyword]

    for alias in aliases:
        for kw in keywords:
            query = f"{city} {alias} {kw}"
            log(f"[INFO] 지역검색: {query}")
            for item in search_local(session, headers, query, log=log):
                key = (item["name"], item["address"])
                if key not in seen:
                    seen[key] = item
            if len(seen) >= max_count:
                return list(seen.values())[:max_count]
    return list(seen.values())[:max_count]


def enrich_business(session, headers, biz, log=print):
    posts = search_blog(session, headers, f"{biz.name} {biz.region}", log=log)
    for post in posts:
        combined = f"{post['title']} {post['desc']}"
        _apply_text(biz, combined, post["link"])
        if biz.instagram and biz.first_visit_event:
            break
        full_text = fetch_blog_text(session, post["link"], log=log)
        _apply_text(biz, full_text, post["link"])
        if biz.instagram and biz.first_visit_event:
            break
    return biz


def _apply_text(biz, text, source_link):
    if not text:
        return
    if not biz.instagram:
        m = INSTAGRAM_RE.search(text)
        if m and "naver" not in m.group(1).lower():
            biz.instagram = "@" + m.group(1)
    if not biz.kakao_channel:
        m = KAKAO_RE.search(text)
        if m:
            biz.kakao_channel = m.group(0)
    if not biz.first_visit_event and "첫방문" in text:
        idx = text.find("첫방문")
        start = text.rfind(" ", 0, max(0, idx - 15)) + 1
        biz.first_visit_event = text[start: idx + 60].replace("\n", " ").strip()
        biz.source_blog = source_link


def crawl(regions, keyword, max_per_region, city="서울", log=print):
    client_id, client_secret = get_credentials()
    headers = {"X-Naver-Client-Id": client_id, "X-Naver-Client-Secret": client_secret}
    session = requests.Session()

    all_businesses = []
    for region in regions:
        log(f"[INFO] === {region} 수집 시작 ===")
        raw_items = collect_region(session, headers, city, region, keyword, max_per_region, log=log)
        log(f"[INFO]  -> {len(raw_items)}곳 발견, 컨택포인트 보강 중...")
        for item in raw_items:
            biz = Business(region=region, name=item["name"], category=item["category"],
                            address=item["address"], phone=item["phone"])
            enrich_business(session, headers, biz, log=log)
            log(f"    - {biz.name} | {biz.address} | IG:{biz.instagram or '-'}")
            all_businesses.append(biz)
    return all_businesses


def save_excel(businesses, output_path, log=print):
    wb = Workbook()
    ws = wb.active
    ws.title = "업체 리스트"
    headers = ["지역", "업체명", "카테고리", "주소", "전화번호", "인스타그램",
               "카카오톡채널", "홈페이지", "첫방문이벤트", "출처(블로그)"]
    ws.append(headers)

    seen = set()
    for b in businesses:
        key = (b.name, b.address)
        if key in seen:
            continue
        seen.add(key)
        ws.append([b.region, b.name, b.category, b.address, b.phone, b.instagram,
                   b.kakao_channel, b.homepage, b.first_visit_event, b.source_blog])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    wb.save(output_path)
    log(f"[INFO] 엑셀 저장 완료: {output_path} (총 {len(seen)}곳)")


def main():
    parser = argparse.ArgumentParser(description="네이버 오픈 API 기반 업체 크롤러")
    parser.add_argument("--city", default="서울", choices=sorted(REGION_ALIASES.keys()),
                         help="지역 별칭 사전을 선택할 도시 (예: 서울, 부산)")
    parser.add_argument("--regions", nargs="+", default=None,
                         help="생략하면 --city의 기본 지역 목록을 사용한다")
    parser.add_argument("--keyword", default="에스테틱")
    parser.add_argument("--max-per-region", type=int, default=15)
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    regions = args.regions or DEFAULT_REGIONS.get(args.city, [args.city])
    output = args.output or f"{args.city}_{args.keyword}_list.xlsx"

    businesses = crawl(regions, args.keyword, args.max_per_region, city=args.city)
    save_excel(businesses, output)


if __name__ == "__main__":
    main()
