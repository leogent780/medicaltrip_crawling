import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Medical Intent Volume"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
BORDER = Border(*(Side(style="thin", color="B0B0B0"),) * 4)
THICK_TOP = Border(top=Side(style="medium", color="000000"))

# ---- Legend ----
ws["A1"] = "SELLE — 의료미용 연결 검색어 볼륨 계산기"
ws["A1"].font = Font(name=FONT, bold=True, size=14)
ws.merge_cells("A1:L1")

legend = [
    ("노란색 셀 (E~I열)", "여기에 KeywordTool.io / Google Keyword Planner에서 확인한 월간 검색량을 직접 입력하세요."),
    ("연한 주황 셀 (D열)", "SERP 실측 기반 Medical Intent % (2026-08-17 조사). 근거 없는 키워드는 비워둠 — 직접 조사 후 입력 권장."),
    ("파란 셀 (J~K열 하단)", "자동 계산되는 합계 (수식, 직접 입력 금지)."),
]
r = 2
for label, desc in legend:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, italic=True, size=9, bold=True)
    ws.cell(row=r, column=2, value=desc).font = Font(name=FONT, italic=True, size=9)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    r += 1

header_row = r + 1
headers = [
    "#", "Keyword", "Category", "Medical Intent %\n(SERP 실측)",
    "Volume - US", "Volume - SG", "Volume - AU", "Volume - CA", "Volume - UK",
    "Total Volume\n(auto)", "Effective Medical\nVolume (auto)", "Notes / Source"
]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = BORDER

# keyword, category, medical_intent (or None), note
rows_data = [
    ("aesthetic clinic seoul", "Bridge", 1.00, "SERP 실측 (2026-08-17): 결과 8/8 클리닉"),
    ("dermatologist seoul", "Bridge", 1.00, "SERP 실측: 결과 9/9 dermatology/clinic"),
    ("best skin clinic seoul", "Bridge", 1.00, "SERP 실측: 6/6 클리닉, foreigner guide 다수"),
    ("skin clinic seoul", "Bridge", 0.80, "SERP 실측: 8/10 클리닉"),
    ("medical tourism korea", "Bridge", 0.78, "SERP 실측: 브로커/정부 채널 다수"),
    ("korean aesthetic treatments", "Bridge", 0.71, "SERP 실측: 학술+클리닉 혼재"),
    ("glass skin treatment", "Bridge", 1.00, "SERP 실측: 6/7 클리닉 (glass skin 단독 대비 극적 반전)"),
    ("korean glass skin treatment", "Bridge", 0.57, "SERP 실측: 클리닉 4/7"),
    ("rejuran korea", "Bottom Funnel", 0.43, "SERP 실측: 제품 판매/클리닉 혼재"),
    ("botox korea", "Bottom Funnel", 0.38, "SERP 실측: 브로커/도매 혼재"),
    ("rejuran korea price", "Bottom Funnel", None, "SERP 미측정 — 직접 확인 필요"),
    ("botox korea price", "Bottom Funnel", None, "SERP 미측정"),
    ("water glow injection seoul", "Bottom Funnel", None, "커뮤니티 언어채굴로 실사용 확인, SERP 미측정"),
    ("ultherapy korea", "Bottom Funnel", None, "SERP 미측정"),
    ("thermage korea", "Bottom Funnel", None, "SERP 미측정"),
    ("skin booster korea", "Bottom Funnel", None, "SERP 미측정"),
    ("juvelook korea", "Bottom Funnel", None, "SERP 미측정"),
    ("cinderella injection korea", "Bottom Funnel", None, "SERP 미측정, 별칭 브랜드"),
    ("acne treatment korea", "Bottom Funnel", None, "SERP 미측정"),
    ("pigmentation treatment korea", "Bottom Funnel", None, "SERP 미측정"),
    ("v line injection korea", "Bottom Funnel", None, "SERP 미측정"),
    ("english speaking dermatologist seoul", "Bridge", None, "SERP 미측정, 커뮤니티에서 반복 등장하는 필터 조건"),
    ("english speaking clinic seoul gangnam", "Bridge", None, "SERP 미측정"),
    ("korean skin clinic for foreigners", "Bridge", None, "경쟁사 콘텐츠 패턴과 일치, SERP 미측정"),
    ("is botox cheaper in korea", "Bridge", None, "실사용 질문형, SERP 미측정"),
    ("korean skin treatment itinerary", "Bridge", None, "경쟁사 블로그 패턴, SERP 미측정"),
    ("best dermatologist seoul foreigners", "Bottom Funnel", None, "SERP 미측정"),
    ("book skin clinic seoul", "Bottom Funnel", None, "SERP 미측정"),
    ("gangnam unni", "Competitor Brand", None, "직접 타깃 키워드 아님, 참고용(경쟁자 흡수 수요)"),
    ("korea skin clinic price", "Bottom Funnel", None, "SERP 미측정"),
    ("korean skincare treatments", "Mother", 0.13, "SERP 실측 (원 제보 키워드)"),
    ("korean skin treatments", "Mother", 0.17, "SERP 실측"),
    ("korean facial", "Mother", 0.20, "SERP 실측 (미국 로컬스파 혼입)"),
    ("k beauty", "Mother", 0.00, "SERP 실측: 전부 제품/문화 콘텐츠"),
    ("korean skincare", "Mother", 0.00, "SERP 실측: 전부 제품 판매"),
]

first_data_row = header_row + 1
for i, (kw, cat, intent, note) in enumerate(rows_data):
    row = first_data_row + i
    ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=2, value=kw).font = Font(name=FONT, size=10)
    ws.cell(row=row, column=3, value=cat).font = Font(name=FONT, size=10)

    d = ws.cell(row=row, column=4)
    if intent is not None:
        d.value = intent
        d.number_format = "0%"
        d.fill = ASSUMPTION_FILL
    d.font = Font(name=FONT, size=10)
    d.alignment = Alignment(horizontal="center")

    for col in range(5, 10):  # E..I volume inputs
        cell = ws.cell(row=row, column=col)
        cell.fill = INPUT_FILL
        cell.number_format = "#,##0"
        cell.font = Font(name=FONT, size=10)

    total_cell = ws.cell(row=row, column=10)
    total_cell.value = f"=SUM(E{row}:I{row})"
    total_cell.number_format = "#,##0"
    total_cell.font = Font(name=FONT, size=10, bold=True)

    eff_cell = ws.cell(row=row, column=11)
    eff_cell.value = f"=IF(D{row}=\"\",\"\",J{row}*D{row})"
    eff_cell.number_format = "#,##0"
    eff_cell.font = Font(name=FONT, size=10, bold=True)

    note_cell = ws.cell(row=row, column=12)
    note_cell.value = note
    note_cell.font = Font(name=FONT, size=9, italic=True)
    note_cell.alignment = Alignment(wrap_text=True)

    for col in range(1, 13):
        ws.cell(row=row, column=col).border = BORDER

last_data_row = first_data_row + len(rows_data) - 1
n_keywords = len(rows_data)
rng_j = f"J{first_data_row}:J{last_data_row}"
rng_k = f"K{first_data_row}:K{last_data_row}"
rng_c = f"C{first_data_row}:C{last_data_row}"

# ---- Headline answer: Bridge + Bottom Funnel only (excludes Mother & Competitor Brand) ----
headline_row = last_data_row + 2
hl = ws.cell(row=headline_row, column=2,
             value="★ 실질적으로 의료미용에 연결되는 검색어 합계 (Bridge + Bottom Funnel만, Mother/경쟁사 브랜드 제외)")
hl.font = Font(name=FONT, bold=True, size=11, color="9C0006")
ws.merge_cells(start_row=headline_row, start_column=2, end_row=headline_row, end_column=8)

ws.cell(row=headline_row + 1, column=9, value="Total Raw Volume").font = Font(name=FONT, bold=True, size=10)
ws.cell(row=headline_row + 1, column=9).alignment = Alignment(horizontal="right")
hl_tv = ws.cell(row=headline_row + 1, column=10,
                 value=f'=SUMIFS({rng_j},{rng_c},"Bridge")+SUMIFS({rng_j},{rng_c},"Bottom Funnel")')
hl_tv.number_format = "#,##0"
hl_tv.font = Font(name=FONT, bold=True, size=12)
hl_tv.fill = TOTAL_FILL
hl_tv.border = THICK_TOP

ws.cell(row=headline_row + 1, column=11,
        value=f'=SUMIFS({rng_k},{rng_c},"Bridge")+SUMIFS({rng_k},{rng_c},"Bottom Funnel")')
hl_emv = ws.cell(row=headline_row + 1, column=11)
hl_emv.number_format = "#,##0"
hl_emv.font = Font(name=FONT, bold=True, size=12)
hl_emv.fill = TOTAL_FILL
hl_emv.border = THICK_TOP
ws.cell(row=headline_row, column=11, value="Effective Medical Volume\n(Volume × Medical Intent%)").font = Font(name=FONT, bold=True, size=9)
ws.cell(row=headline_row, column=11).alignment = Alignment(horizontal="right", wrap_text=True)

note = ws.cell(row=headline_row + 2, column=2,
               value="주: Effective Medical Volume은 Medical Intent %가 비어있는(SERP 미측정) 키워드는 자동으로 0으로 계산되어 과소평가됩니다 — D열을 채울수록 정확해집니다.")
note.font = Font(name=FONT, italic=True, size=8)
ws.merge_cells(start_row=headline_row + 2, start_column=2, end_row=headline_row + 2, end_column=11)

# ---- Reference: full-table sum incl. Mother + Competitor Brand ----
summary_row = headline_row + 4
ws.cell(row=summary_row, column=2,
        value=f"참고: 전체 {n_keywords}개 키워드 합계 (Mother/경쟁사 브랜드 포함, 참고용— 위 헤드라인 숫자를 사용할 것)").font = Font(name=FONT, italic=True, size=9)
ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=8)

ws.cell(row=summary_row, column=9, value="Total Raw Volume").font = Font(name=FONT, size=9)
ws.cell(row=summary_row, column=9).alignment = Alignment(horizontal="right")
tv = ws.cell(row=summary_row, column=10, value=f"=SUM({rng_j})")
tv.number_format = "#,##0"
tv.font = Font(name=FONT, size=10)

emv = ws.cell(row=summary_row, column=11, value=f"=SUM({rng_k})")
emv.number_format = "#,##0"
emv.font = Font(name=FONT, size=10)

# category breakdown using SUMIFS
cat_row = summary_row + 2
ws.cell(row=cat_row, column=2, value="카테고리별 합계").font = Font(name=FONT, bold=True, size=10)
cat_row += 1
cats = ["Mother", "Bridge", "Bottom Funnel", "Competitor Brand"]
ws.cell(row=cat_row, column=2, value="Category").font = Font(name=FONT, bold=True, size=9)
ws.cell(row=cat_row, column=3, value="Total Volume").font = Font(name=FONT, bold=True, size=9)
ws.cell(row=cat_row, column=4, value="Effective Medical Volume").font = Font(name=FONT, bold=True, size=9)
for i, cat in enumerate(cats):
    rr = cat_row + 1 + i
    ws.cell(row=rr, column=2, value=cat).font = Font(name=FONT, size=9)
    tvc = ws.cell(row=rr, column=3, value=f'=SUMIFS({rng_j},{rng_c},"{cat}")')
    tvc.number_format = "#,##0"
    emvc = ws.cell(row=rr, column=4, value=f'=SUMIFS({rng_k},{rng_c},"{cat}")')
    emvc.number_format = "#,##0"

# column widths
widths = {1: 4, 2: 30, 3: 16, 4: 14, 5: 11, 6: 11, 7: 11, 8: 11, 9: 11, 10: 13, 11: 15, 12: 40}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = f"E{first_data_row}"

wb.save("/home/user/medicaltrip_crawling/research/en_medical_intent_volume_calculator.xlsx")
print("saved")
